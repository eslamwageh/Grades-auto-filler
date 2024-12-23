from commonfunctions import *
from joblib import dump, load
import pytesseract
from openpyxl.styles import PatternFill

def get_edges(image):
    blurred = cv2.GaussianBlur(image, (21, 21), 0.7)
    gray = cv2.cvtColor(blurred, cv2.COLOR_RGB2GRAY)
    
    edges = cv2.Canny(gray,50, 80)

    # show_images([image, gray, blurred, edges], ['img', 'gray', 'blurred', 'edges'])

    return edges

def getPerspective(img):
    imgContours = img.copy()
    imgCorners = img.copy()
    
    
    edges = get_edges(img)

    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (10, 10))

    closed_edges = cv2.morphologyEx(edges, cv2.MORPH_CLOSE, kernel, iterations=1)

    contours, _ = cv2.findContours(closed_edges, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)


    rectContours = rectContour(contours)

    cv2.drawContours(imgContours, rectContours, -1, (0, 255, 0), 2)

    thirdBiggestRect = getCornerPoints(rectContours[2])

    cv2.drawContours(imgCorners, thirdBiggestRect, -1, (0, 255, 0), 50)

    thirdBiggestRect = reorder(thirdBiggestRect)

    pt1 = np.float32(thirdBiggestRect)
    pt2 = np.float32([[0, 0], [img.shape[1], 0], [0, img.shape[0]], [img.shape[1], img.shape[0]]])

    matrix = cv2.getPerspectiveTransform(pt1, pt2)
    # Apply the perspective transformation to the image
    warped_img = cv2.warpPerspective(img, matrix, (img.shape[1], img.shape[0]))


    # SHOWING THE IMAGES FOR CLARITY 
    # show_images([closed_edges], ['closed_edges'])
    # show_images([imgContours,imgCorners], ['imgContours','imgCorners'])
    # show_images([warped_img], ["warped"])
    return warped_img

def getLines(full_paper):
    edges = get_edges(full_paper)

    edges = borderTheImage(edges)

    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (5, 5))

    closed_edges = cv2.morphologyEx(edges, cv2.MORPH_CLOSE, kernel, iterations=6)

    # Detect lines using the Hough Line Transform
    lines = cv2.HoughLinesP(closed_edges, rho=1, theta=np.pi/180, threshold=100, minLineLength=2000, maxLineGap=50)

    horizontal_lines = []
    vertical_lines = []

    for line in lines:
        x1, y1, x2, y2 = line[0]
        if abs(y1 - y2) < 10:  # Horizontal line
            horizontal_lines.append(line[0])
        elif abs(x1 - x2) < 10:  # Vertical line
            vertical_lines.append(line[0])

    clustered_horizontal = cluster_lines(full_paper, horizontal_lines, orientation='horizontal', threshold=10)
    clustered_vertical = cluster_lines(full_paper, vertical_lines, orientation='vertical', threshold=10)
    clustered_lines = clustered_horizontal + clustered_vertical



    # show_images([closed_edges], ["closed edges"])

    # showLines(full_paper, horizontal_lines, vertical_lines, lines, 0)
    # showLines(full_paper, clustered_horizontal, clustered_vertical, clustered_lines, 1)

    return clustered_horizontal, clustered_vertical

def showLines(full_paper, horizontal_lines, vertical_lines, lines, clustered = 0):
    horizontal = full_paper.copy()
    vertical = full_paper.copy()
    output = full_paper.copy()

    if horizontal_lines is not None:
        for line in horizontal_lines:
            x1, y1, x2, y2 = line
            cv2.line(horizontal, (x1, y1), (x2, y2), (0, 255, 0), 10)

    if vertical_lines is not None:
        for line in vertical_lines:
            x1, y1, x2, y2 = line
            cv2.line(vertical, (x1, y1), (x2, y2), (0, 255, 0), 10)

    # Draw detected lines
    if lines is not None:
        for line in lines:
            if (clustered == 0):
                x1, y1, x2, y2 = line[0]
            else:
                x1, y1, x2, y2 = line
            cv2.line(output, (x1, y1), (x2, y2), (0, 255, 0), 10)
    # if (clustered == 0):
    #     show_images([horizontal, vertical, output], ["horizontal", "vertical", "lines"])
    # else:
    #     show_images([horizontal, vertical, output], ["clustered horizontal", "clustered vertical", "clustered lines"])

def borderTheImage(image, top = 3, bottom = 3, right = 5, left = 2):
    # Fill the borders with ones
    image[:top, :] = 255  # Top border
    image[-bottom:, :] = 255  # Bottom border
    image[:, :left] = 255  # Left border
    image[:, -right:] = 255  # Right border
    return image

def cluster_lines(full_paper, lines, orientation='horizontal', threshold=10, shifting = 2):
    """
    Cluster lines that are close to each other.

    Parameters:
    - lines: list of lines [x1, y1, x2, y2]
    - orientation: 'horizontal' or 'vertical'
    - threshold: distance threshold for clustering

    Returns:
    - clustered_lines: list of clustered lines
    """
    if orientation == 'horizontal':
        # Use the y-coordinate for clustering
        lines_sorted = sorted(lines, key=lambda line: line[1])
    else:
        # Use the x-coordinate for clustering
        lines_sorted = sorted(lines, key=lambda line: line[0])
    
    clustered_lines = []
    cluster = []
    
    for line in lines_sorted:
        if not cluster:
            cluster.append(line)
            continue
        
        if orientation == 'horizontal':
            # Compare y-coordinates
            if abs(line[1] - cluster[-1][1]) < threshold:
                cluster.append(line)
            else:
                # Merge the current cluster
                avg_y = int(np.mean([l[1] for l in cluster])) - shifting
                x_start = min(l[0] for l in cluster)
                x_end = max(l[2] for l in cluster)
                # clustered_lines.append([x_start, avg_y, x_end, avg_y])
                clustered_lines.append([0, avg_y, full_paper.shape[1] - 1, avg_y])
                cluster = [line]
        else:
            # Compare x-coordinates
            if abs(line[0] - cluster[-1][0]) < threshold:
                cluster.append(line)
            else:
                # Merge the current cluster
                avg_x = int(np.mean([l[0] for l in cluster])) - shifting
                y_start = min(l[1] for l in cluster)
                y_end = max(l[3] for l in cluster)
                # clustered_lines.append([avg_x, y_start, avg_x, y_end])
                clustered_lines.append([avg_x, 0, avg_x, full_paper.shape[0] - 1])
                cluster = [line]
    
    # Merge the last cluster
    if cluster:
        if orientation == 'horizontal':
            avg_y = int(np.mean([l[1] for l in cluster])) - shifting
            x_start = min(l[0] for l in cluster)
            x_end = max(l[2] for l in cluster)
            # clustered_lines.append([x_start, avg_y, x_end, avg_y])
            clustered_lines.append([0, avg_y, full_paper.shape[1] - 1, avg_y])
        else:
            avg_x = int(np.mean([l[0] for l in cluster])) - shifting
            y_start = min(l[1] for l in cluster)
            y_end = max(l[3] for l in cluster)
            # clustered_lines.append([avg_x, y_start, avg_x, y_end])
            clustered_lines.append([avg_x, 0, avg_x, full_paper.shape[0] - 1])
    
    return clustered_lines

def extractCells(full_paper, clustered_horizontal, clustered_vertical):
    # Sort the horizontal and vertical lines
    clustered_horizontal = sorted(clustered_horizontal, key=lambda line: line[1])  # Sort by y
    clustered_vertical = sorted(clustered_vertical, key=lambda line: line[0])      # Sort by x

    # Number of rows and columns
    num_rows = len(clustered_horizontal) - 1
    num_cols = len(clustered_vertical) - 1

    print(f"Table has {num_rows} rows and {num_cols} columns.")

    cells = []

    for i in range(num_rows):
        row_cells = []
        for j in range(num_cols):
            # Top-left corner
            x1 = clustered_vertical[j][0]
            y1 = clustered_horizontal[i][1]
            
            # Bottom-right corner
            x2 = clustered_vertical[j + 1][0]
            y2 = clustered_horizontal[i + 1][1]
            
            # Crop the cell from the original image
            cell = full_paper[y1:y2, x1:x2]
            row_cells.append(cell)
            #show_images([cell], [f"{i}, {j}"])
        cells.append(row_cells)
    
    print(f"Extracted {len(cells) * len(cells[0])} cells.")
    return cells
    

def predictCells(cells, digits_models, symbols_models, selected_method, sheet):
    # Define red fill
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for row_cells in cells:
        result_id = ""
        answers = []
        for i in range(len(row_cells)):
            height, width = row_cells[i].shape[:2]

            if (i == 0):
                result_id = predictID(row_cells[i], selected_method, digits_models)
            elif (i == 3):
                start_x = (width - 140) // 2
                start_y = (height - 125) // 2
                end_x = start_x + 120
                end_y = start_y + 125

                # Crop the image
                cropped_cell = row_cells[i][start_y:end_y, start_x:end_x]
                # show_images([cropped_cell],["cropped cell"])
                current_answer = predict_digit(cropped_cell, digits_models, selected_method)
                answers.append(current_answer)

                print(f"The digit predicted is: {predict_digit(cropped_cell, digits_models, selected_method)}")
            elif (i > 3):
                start_x = (width - 250) // 2
                start_y = (height - 150) // 2
                end_x = start_x + 250
                end_y = start_y + 150

                # Crop the image
                cropped_cell = row_cells[i][start_y:end_y, start_x:end_x]
                # show_images([cropped_cell],["cropped cell"])

                result = None
                symbol = predict_symbol(cropped_cell, symbols_models)
                match symbol:
                    case 'check':
                        result = 5
                    case 'empty':
                        result = ""
                    case 'vertical1':
                        result = 1
                    case 'vertical2':
                        result = 2
                    case 'vertical3':
                        result = 3
                    case 'vertical4':
                        result = 4
                    case 'vertical5':
                        result = 5
                    case 'horizontal1':
                        result = 4
                    case 'horizontal2':
                        result = 3
                    case 'horizontal3':
                        result = 2
                    case 'horizontal4':
                        result = 1
                    case 'question':
                        result = '%'
                    case 'square':
                        result = 0
                    case _:
                        result = None
                answers.append(result)

                print(f"The symbol predicted is: {symbol}")
            
        sheet.append([result_id] + answers)
        for col_index, answer in enumerate(answers, start=2):  # Starting from the 2nd column
            if '%' in str(answer):  # Check if the answer contains '%'
                cell = sheet.cell(row=sheet.max_row, column=col_index)
                cell.fill = red_fill  # Set background color to red
                cell.value = ""  # Set cell value to empty


def preprocessIDDigit(digit_img):
    # Find the brightest color in the image
    brightest_color = digit_img[0, 0]  # Max across rows and columns for each channel

    # Create a new blank image of size (70, 70) filled with the brightest color
    result_img = np.full((70, 70, 3), brightest_color, dtype=np.uint8)

    # Resize the digit image while maintaining its aspect ratio
    digit_h, digit_w = digit_img.shape[:2]
    scale = min(70 / digit_h, 70 / digit_w)
    new_w, new_h = int(digit_w * scale), int(digit_h * scale)
    resized_digit = cv2.resize(digit_img, (new_w, new_h))

    # Calculate the offset to center the digit
    x_offset = (70 - new_w) // 2
    y_offset = (70 - new_h) // 2

    # Place the resized digit in the center of the result image
    result_img[y_offset:y_offset+new_h, x_offset:x_offset+new_w] = resized_digit
    return result_img

def predictID(img, selected_method, digits_models):
    result = ""
    if (selected_method == "OCR"):
        result = ocr_pytesseract_number_extraction_default(img)
        print ("OCR is not installed :(")
        print(f"The ID predicted is: {result}")

    else:
        # show_images([img], ["img"])
        id_contours_img = img.copy()

        blurred = cv2.GaussianBlur(img, (11, 11), 2)
        gray = cv2.cvtColor(blurred, cv2.COLOR_RGB2GRAY)
        
        edges = cv2.Canny(gray,60, 100)

        edges[ : 15, :] = 0
        edges[-15: , :] = 0
        edges[:, :15] = 0
        edges[:, -15:] = 0

        kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 100))
        closed_edges = cv2.morphologyEx(edges, cv2.MORPH_CLOSE, kernel, iterations=1)

        # show_images([img, gray, blurred, edges, closed_edges], ['img', 'gray', 'blurred', 'edges', 'closed_edges'])

        # edges = get_edges(img)
        # edges = borderTheImage(edges, 12, 12, 12, 12)

        
        id_contours, _ = cv2.findContours(closed_edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
        # paper_rectContours = rectContour(paper_contours)
        cv2.drawContours(id_contours_img, id_contours, -1, (0, 255, 0), 1)

        id_contours = sorted(id_contours, key=lambda contour: cv2.boundingRect(contour)[0])

        show_images([id_contours_img], ["id contours"])

        for i, contour in enumerate(id_contours):
            x, y, w, h = cv2.boundingRect(contour)

            if (h < 1.5 * w):
                digit_img1 = img[y:y+h, x:x+w//2]
                digit_img2 = img[y:y+h, x+w//2:x+w]

                result_img = preprocessIDDigit(digit_img1)
                predicted_digit = predict_digit(result_img, digits_models, selected_method)
                if predicted_digit == 10:
                    predicted_digit = 0
                print(f"The digit predicted is: {predicted_digit}")
                result += str(predicted_digit)

                result_img = preprocessIDDigit(digit_img2)
                predicted_digit = predict_digit(result_img, digits_models, selected_method)
                if predicted_digit == 10:
                    predicted_digit = 0
                print(f"The digit predicted is: {predicted_digit}")
                result += str(predicted_digit)

            else :
                digit_img1 = img[y:y+h, x:x+w]

                result_img = preprocessIDDigit(digit_img1)
                predicted_digit = predict_digit(result_img, digits_models, selected_method)
                if predicted_digit == 10:
                    predicted_digit = 0
                print(f"The digit predicted is: {predicted_digit}")
                result += str(predicted_digit)
            
            # Display the resulting image
            # show_images([result_img], [f"Digit {i+1}"])
            # show_images([digit_img], [f"Digit {i+1}"])
    return result


############################################# CLASSIFICATON FUNCTIONS ##################################

def extract_hog_features(img):
    """
    Extracts Histogram of Oriented Gradients (HOG) features from the input image.
    This involves resizing the image to a fixed size, dividing it into cells and blocks, 
    computing gradient histograms, and flattening the result into a feature vector. 
    The extracted features are useful for machine learning tasks like image classification and detection.
    """
    target_img_size = (128, 128)
    img = cv2.resize(img, dsize=target_img_size)
    win_size = (128, 128) 
    cell_size = (16, 16) #Divides the window into smaller cells (4x4 pixels per cell).
    block_size_in_cells = (8, 8) 

    # divides window to blocks then blocks to cells
    block_size = (block_size_in_cells[1] * cell_size[1], block_size_in_cells[0] * cell_size[0])
    block_stride = (cell_size[1], cell_size[0]) #Determines how much the block moves at each step (4x4 pixels, matching cell size).
    nbins = 9  # Number of orientation bins
    hog = cv2.HOGDescriptor(win_size, block_size, block_stride, cell_size, nbins)
    h = hog.compute(img)
    h = h.flatten()
    return h.flatten() #This array is used as input for machine learning models.

def load_model():
    # Load the digits models dictionary
    loaded_digits_models = load('digits_models.joblib')
    print('Loaded digits models from digits_models.joblib')
    print(loaded_digits_models)

    # Load the symbols models dictionary
    loaded_symbols_models = load('symbols_models.joblib')
    print('Loaded symbols models from symbols_models.joblib')

    # Return the loaded models
    return loaded_digits_models, loaded_symbols_models
    

    
def predict_symbol(img, symbols_models):

    test_features=extract_hog_features(img)
    predicted_symbol=symbols_models['SVM'].predict([test_features])
    return predicted_symbol



def predict_digit(img, digits_models, selected_method):
    if (selected_method == "OCR"):
        print ("OCR is not installed :(")
        return ocr_pytesseract_number_extraction(img);
    else:
        test_features=extract_hog_features(img)
        predicted_digit=digits_models['SVM'].predict([test_features])
        predicted_digit_num = ord(predicted_digit[0].lower()) - ord('a') + 1
    
    return predicted_digit_num




########################################## OCR FUNCTIONS ##############################################


def ocr_pytesseract_number_extraction_default(image):
    # Open the image using Pillow
    #you can remove the config to detect the text if you want but we only using it for digits detection
    extracted_text = pytesseract.image_to_string(image, config='--psm 6 -c tessedit_char_whitelist=0123456789')
    return extracted_text

def ocr_pytesseract_number_extraction(img):
    if img is None:
        print("Error: Unable to load the image.")
        return None

    # Ensure the image is in grayscale format
    if len(img.shape) == 3 and img.shape[2] == 3:
        # Convert color image to grayscale
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    elif len(img.shape) == 2:
        # Already in grayscale, no need to convert
        gray = img
    else:
        print("Error: Invalid image format.")
        return None


    # Apply thresholding to the grayscale image to improve OCR accuracy for images with inconsistent lighting or low contrast.
    _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)

    # custom configurations for single digits outputs
    # --psm 10 (for single character recognition)
    custom_config = r'--oem 3 --psm 10 outputbase digits'
    text = pytesseract.image_to_string(binary, config=custom_config)

    return text


# def find_intersections(horizontal, vertical):
#     """
#     Find intersection points between horizontal and vertical lines.

#     Parameters:
#     - horizontal: list of horizontal lines [x1, y, x2, y]
#     - vertical: list of vertical lines [x, y1, x, y2]

#     Returns:
#     - intersections: list of (x, y) tuples
#     """
#     intersections = []
#     for h_line in horizontal:
#         _, y, _, _ = h_line
#         for v_line in vertical:
#             x, _, _, _ = v_line
#             intersections.append((x, y))
#     return intersections


# def getID(full_paper):
#     full_paper_contours = full_paper.copy()
#     full_paper_largest_contour = full_paper.copy()

#     gray = cv2.cvtColor(full_paper, cv2.COLOR_RGB2GRAY)
#     blurred = cv2.GaussianBlur(gray, (3, 3), 0.7)
#     edges = cv2.Canny(blurred, 50, 120)


#     # Create a mask with zeros at the borders and ones elsewhere
#     height,width = edges.shape
#     border_mask = np.zeros((height, width), dtype=np.uint8)
#     border_mask[10:-10,10:-10] = 1  # Leave a 1-pixel border untouched

#     # Dilate the edges
#     print(edges.shape)
#     kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 5))
#     vertically_closed = cv2.morphologyEx(edges, cv2.MORPH_CLOSE, kernel, iterations=10)
#     show_images([vertically_closed], ["vertically_closed"])


#     kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (8, 5))
#     closed = cv2.morphologyEx(vertically_closed, cv2.MORPH_CLOSE, kernel, iterations=3)



#     # # Apply the border mask to retain original borders
#     closed_paper = np.where(border_mask, closed, edges)

#     closed_paper[:10,:] = 0
#     closed_paper[-10:, :] = 0
#     closed_paper[:, :10] = 0
#     closed_paper[:, -10:] = 0
#     paper_contours, _ = cv2.findContours(closed_paper, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
#     # paper_rectContours = rectContour(paper_contours)
#     cv2.drawContours(full_paper_contours, paper_contours, -1, (0, 255, 0), 3)
#     # paper_largest_contour = paper_rectContours[0]
#     paper_largest_contour = getLowerBiggestContour(paper_contours)
#     cv2.drawContours(full_paper_largest_contour, [paper_largest_contour], -1, (0, 255, 0), 3)


#     x, y, w, h = cv2.boundingRect(paper_largest_contour)
#     questions = full_paper[y:y+h, x:x+w] 

#     # SHOWING THE IMAGES FOR CLARITY 

#     show_images([edges, closed_paper, full_paper_contours, full_paper_largest_contour, questions], ["edges","closed", "contours", "largest_contour", "questions"])
#     return questions